#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb  5 21:47:56 2021

@author: bernhard

"""

import numpy as np
from datetime import date
from google_trans_new import google_translator
import easygui
import openpyxl
from openpyxl.styles.borders import Border, Side
import os
from nacres_from_thorlabs import nacres_from_thorlabs

# define cell boarders join appropriate cells
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def num2str(num):
    output = f'{num:.2f}'.replace('.',',')
    if len(output) > 6:
        pos = len(output) - 6
        output = output[:pos] + '.' + output[pos:]
    return output

# INPUT
file_path = easygui.fileopenbox("select a shoppingCart exel file created from Thorlabs site")
target_dir = os.sep.join(file_path.split(os.sep)[:-1])+os.sep

discount = float(easygui.integerbox("Enter the discount in %", "Discount", 9))/100.
shipping_cost = float(easygui.enterbox("Enter the shipping cost HT"))


# load input data - check first the existence of .xls (default from thorlabs site), 
# then the .xlsx (from original implementation of this code)
if file_path.split('.')[-1] == 'xls':
    # load with pandas if old .xls format
    import pandas as pd
    pd_wb = pd.read_excel(file_path)
    input_data = pd_wb.values.tolist()
elif file_path.split('.')[-1] == 'xlsx':
    # load with openpyxl if .xlsx
    input_wb = openpyxl.load_workbook(file_path)
    input_ws = input_wb.active
    input_data = list( input_ws.values )[1:]

# handle som weird stuffs with xlsx...
for idx in range(len(input_data)):
    if input_data[idx][0] is None:
        idx = idx-1
        break
nr_items = idx+1
print(f'Processing {nr_items} items from {file_path}')

# open output file
output_wb = openpyxl.load_workbook('thorlabsBC_template.xlsx')
output_ws = output_wb.active

# set date of today
today = date.today()
output_ws.cell(row=3, column=10).value = today.strftime("%d/%m/%Y")

# insert necessary rows in output file
first_row = 10;
output_ws.insert_rows(first_row, amount=(nr_items-1))  
# output_ws.insert_rows(first_row, amount=(nr_items))  

# loop through items
price_unit = np.zeros((nr_items,))
price_unit_disc = np.zeros((nr_items,))
price_full_disc = np.zeros((nr_items,))

for idx in range(nr_items):

    # get values from input list
    desc = input_data[idx][1]
    quantity = input_data[idx][3]
    price_unit[idx] = input_data[idx][4]
    
    # translate decription to french
    translator = google_translator() 
    desc_fr = translator.translate(desc, lang_tgt='fr')
    
    # calulate untaxed, dicount and full price
    price_unit_disc[idx] = round( (1-discount) * price_unit[idx], 2 )
    price_full_disc[idx] = round( quantity * price_unit_disc[idx], 2 )
    
    # if available from the dict, fill the NACRES code from the thorlabs code
    # input_data[idx][0] is the key, which is processsed through the nacres_from_thorlabs()
    nacres = nacres_from_thorlabs(input_data[idx][0], input_data[idx][1])
    
    # fill output file entries
    output_ws.cell(row=first_row+idx, column=1).value = quantity
    output_ws.cell(row=first_row+idx, column=2).value = desc_fr
    output_ws.cell(row=first_row+idx, column=5).value = input_data[idx][0]
    output_ws.cell(row=first_row+idx, column=7).value = nacres
    output_ws.cell(row=first_row+idx, column=8).value = num2str(price_unit_disc[idx])
    output_ws.cell(row=first_row+idx, column=9).value = num2str(discount*100)
    output_ws.cell(row=first_row+idx, column=10).value = num2str(price_full_disc[idx])
    
    # fix cell boarders join appropriate cells
    for idx_col in range(1,11):
        output_ws.cell(row=first_row+idx, column=idx_col).border = thin_border
    output_ws.merge_cells(start_row=first_row+idx, start_column=2, end_row=first_row+idx, end_column=4)
    output_ws.merge_cells(start_row=first_row+idx, start_column=5, end_row=first_row+idx, end_column=6)

# insert the shipping cost
output_ws.cell(row=first_row+nr_items, column=10).value = num2str(shipping_cost)

# calculate total price    
price_nr_brut_disc_sum = round(np.sum(price_full_disc)+shipping_cost,2)

# insert total amount in output file
output_ws.cell(row=first_row+nr_items+1, column=10).value = num2str(price_nr_brut_disc_sum)
    
# save output file
output_file = target_dir + 'thorlabsBC.xlsx'
output_wb.save(output_file)
print(f'Output written in {output_file}')